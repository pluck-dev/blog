"""
v2: markdown_full(CSV) 기반 MinHash + LSH 클러스터링.
- 빈 본문은 제외
- shingle k=9 (더 엄격), threshold=0.75
"""
import csv, re, time, sys
from collections import defaultdict
from datasketch import MinHash, MinHashLSH
import openpyxl
from openpyxl.styles import Font, PatternFill

csv.field_size_limit(sys.maxsize)

SRC = '/Users/simjaehyeong/Desktop/adrock/seo/data/drivingteacher_blog_posts.csv'
DST = '/Users/simjaehyeong/Desktop/adrock/seo/data/drivingteacher_blog_posts_classified.xlsx'

SHINGLE_K = 9
NUM_PERM = 128
JACCARD_THRESHOLD = 0.75
MIN_LEN = 200  # 본문 최소 길이 (이하는 클러스터링에서 제외)

def normalize(text: str) -> str:
    text = text or ''
    text = re.sub(r'!\[.*?\]\(.*?\)', ' ', text)        # 마크다운 이미지
    text = re.sub(r'\[(.*?)\]\(.*?\)', r'\1', text)     # 마크다운 링크 -> 텍스트만
    text = re.sub(r'https?://\S+', ' ', text)
    text = re.sub(r'[#`*_>~|\-=]+', ' ', text)          # 마크다운 기호
    text = re.sub(r'[^\w가-힣]+', ' ', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text

def shingles(text: str, k: int = SHINGLE_K):
    if len(text) < k:
        return set([text]) if text else set()
    return {text[i:i+k] for i in range(len(text) - k + 1)}

def jaccard(a, b):
    if not a or not b: return 0.0
    return len(a & b) / len(a | b)

def main():
    t0 = time.time()
    print(f'[{time.time()-t0:5.1f}s] CSV 로드 시작...')
    posts = []  # (id, title, url, template, createdAt, content_normalized)
    skipped_empty = 0
    skipped_short = 0
    with open(SRC, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            md = row.get('markdown_full', '') or ''
            if not md:
                skipped_empty += 1
                continue
            norm = normalize(md)
            if len(norm) < MIN_LEN:
                skipped_short += 1
                continue
            posts.append((
                row.get('id',''), row.get('title',''), row.get('url',''),
                row.get('templateLabel',''), row.get('createdAt',''),
                norm
            ))
            if len(posts) % 5000 == 0:
                print(f'[{time.time()-t0:5.1f}s]  로드 {len(posts)}건')
    print(f'[{time.time()-t0:5.1f}s] 로드 완료: {len(posts)}건 / 빈 본문 제외 {skipped_empty} / 짧은 본문 제외 {skipped_short}')

    print(f'[{time.time()-t0:5.1f}s] MinHash 생성 (k={SHINGLE_K}, perm={NUM_PERM})...')
    minhashes = []
    sh_cache = []
    for i, p in enumerate(posts):
        sh = shingles(p[5])
        sh_cache.append(sh)
        m = MinHash(num_perm=NUM_PERM)
        for s in sh:
            m.update(s.encode('utf-8'))
        minhashes.append(m)
        if (i+1) % 5000 == 0:
            print(f'[{time.time()-t0:5.1f}s]   {i+1}/{len(posts)}')

    print(f'[{time.time()-t0:5.1f}s] LSH 인덱스 구축...')
    lsh = MinHashLSH(threshold=JACCARD_THRESHOLD, num_perm=NUM_PERM)
    for i, m in enumerate(minhashes):
        lsh.insert(i, m)

    print(f'[{time.time()-t0:5.1f}s] 후보 쌍 추출...')
    pairs = []
    seen = set()
    for i, m in enumerate(minhashes):
        for j in lsh.query(m):
            if i >= j: continue
            key = (i, j)
            if key in seen: continue
            seen.add(key)
            jsim = jaccard(sh_cache[i], sh_cache[j])
            if jsim >= JACCARD_THRESHOLD:
                pairs.append((i, j, jsim))
        if (i+1) % 5000 == 0:
            print(f'[{time.time()-t0:5.1f}s]   {i+1}/{len(posts)} pairs={len(pairs)}')
    print(f'[{time.time()-t0:5.1f}s] 유사 쌍: {len(pairs)}')

    # Union-Find
    parent = list(range(len(posts)))
    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]; x = parent[x]
        return x
    def union(a,b):
        ra, rb = find(a), find(b)
        if ra != rb: parent[ra] = rb
    for i, j, _ in pairs:
        union(i, j)
    clusters = defaultdict(list)
    for i in range(len(posts)):
        clusters[find(i)].append(i)
    dup_clusters = {k:v for k,v in clusters.items() if len(v) > 1}
    print(f'[{time.time()-t0:5.1f}s] 클러스터(>=2): {len(dup_clusters)}, 영향받는 글: {sum(len(v) for v in dup_clusters.values())}')

    # 분포
    size_dist = defaultdict(int)
    for v in dup_clusters.values():
        size_dist[len(v)] += 1
    print('클러스터 크기 분포:')
    for sz in sorted(size_dist):
        print(f'   size {sz}: {size_dist[sz]}건')

    # 유사도 분포
    sim_bins = defaultdict(int)
    for _, _, s in pairs:
        if s >= 0.95: sim_bins['0.95-1.00'] += 1
        elif s >= 0.85: sim_bins['0.85-0.95'] += 1
        else: sim_bins['0.75-0.85'] += 1
    print('유사도 분포:')
    for k in ['0.95-1.00','0.85-0.95','0.75-0.85']:
        print(f'   {k}: {sim_bins[k]}')

    # 저장
    print(f'[{time.time()-t0:5.1f}s] xlsx 저장...')
    wb = openpyxl.load_workbook(DST)
    for sn in ['Duplicate_Clusters', 'Duplicate_Pairs']:
        if sn in wb.sheetnames: del wb[sn]

    hdr_font = Font(bold=True, color='FFFFFF')
    hdr_fill = PatternFill('solid', fgColor='C00000')

    ws_c = wb.create_sheet('Duplicate_Clusters')
    ws_c.append(['cluster_id','cluster_size','id','title','url','templateLabel','createdAt'])
    for c in ws_c[1]: c.font = hdr_font; c.fill = hdr_fill
    ws_c.freeze_panes = 'A2'
    cid = 0
    for root, members in sorted(dup_clusters.items(), key=lambda x: -len(x[1])):
        cid += 1
        for mi in members:
            p = posts[mi]
            ws_c.append([cid, len(members), p[0], p[1], p[2], p[3], p[4]])
    for col, w in [('A',10),('B',12),('C',22),('D',60),('E',55),('F',30),('G',25)]:
        ws_c.column_dimensions[col].width = w

    ws_p = wb.create_sheet('Duplicate_Pairs')
    ws_p.append(['jaccard','id_a','title_a','url_a','id_b','title_b','url_b','templateLabel_a','templateLabel_b','createdAt_a','createdAt_b'])
    for c in ws_p[1]: c.font = hdr_font; c.fill = hdr_fill
    ws_p.freeze_panes = 'A2'
    for i, j, sim in sorted(pairs, key=lambda x:-x[2]):
        pa, pb = posts[i], posts[j]
        ws_p.append([round(sim,4), pa[0], pa[1], pa[2], pb[0], pb[1], pb[2], pa[3], pb[3], pa[4], pb[4]])
    for col, w in [('A',10),('B',22),('C',55),('D',50),('E',22),('F',55),('G',50),('H',28),('I',28),('J',22),('K',22)]:
        ws_p.column_dimensions[col].width = w

    wb.save(DST)
    print(f'[{time.time()-t0:5.1f}s] 완료: {DST}')

if __name__ == '__main__':
    main()
