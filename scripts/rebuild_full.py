"""
통합 빌드:
1) lite xlsx -> Human/AI_generated/All 시트 (templateLabel 기준)
2) CSV(markdown_full) -> MinHash/LSH 유사 본문 클러스터링
3) 최종 xlsx 한 번에 저장
"""
import csv, re, sys, time
from collections import defaultdict
from datasketch import MinHash, MinHashLSH
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

csv.field_size_limit(sys.maxsize)

LITE = '/Users/simjaehyeong/Desktop/adrock/seo/data/drivingteacher_blog_posts_lite.xlsx'
CSV  = '/Users/simjaehyeong/Desktop/adrock/seo/data/drivingteacher_blog_posts.csv'
DST  = '/Users/simjaehyeong/Desktop/adrock/seo/data/drivingteacher_blog_posts_classified.xlsx'

SHINGLE_K = 9
NUM_PERM = 128
JACCARD_THRESHOLD = 0.75
MIN_LEN = 200

def normalize(text):
    text = text or ''
    text = re.sub(r'!\[.*?\]\(.*?\)', ' ', text)
    text = re.sub(r'\[(.*?)\]\(.*?\)', r'\1', text)
    text = re.sub(r'https?://\S+', ' ', text)
    text = re.sub(r'[#`*_>~|\-=]+', ' ', text)
    text = re.sub(r'[^\w가-힣]+', ' ', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text

def shingles(text, k=SHINGLE_K):
    if len(text) < k:
        return set([text]) if text else set()
    return {text[i:i+k] for i in range(len(text) - k + 1)}

def jaccard(a, b):
    if not a or not b: return 0.0
    return len(a & b) / len(a | b)

def main():
    t0 = time.time()
    print(f'[{time.time()-t0:5.1f}s] === STEP 1: 분류 (lite xlsx) ===')
    src = openpyxl.load_workbook(LITE, read_only=True)
    ws_src = src['Posts']
    rows = list(ws_src.iter_rows(values_only=True))
    H = rows[0]; data = rows[1:]
    idx_lite = {c:i for i,c in enumerate(H)}
    print(f'[{time.time()-t0:5.1f}s]  -> {len(data)}건')

    TPL = idx_lite['templateLabel']
    human = [r for r in data if not r[TPL] or not str(r[TPL]).strip()]
    ai    = [r for r in data if r[TPL] and str(r[TPL]).strip()]
    print(f'[{time.time()-t0:5.1f}s] Human {len(human)} / AI {len(ai)} / All {len(data)}')

    print(f'\n[{time.time()-t0:5.1f}s] === STEP 2: CSV 본문 로드 ===')
    posts = []
    skipped_empty = skipped_short = 0
    with open(CSV, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            md = row.get('markdown_full', '') or ''
            if not md: skipped_empty += 1; continue
            norm = normalize(md)
            if len(norm) < MIN_LEN: skipped_short += 1; continue
            posts.append((
                row.get('id',''), row.get('title',''), row.get('url',''),
                row.get('templateLabel',''), row.get('createdAt',''),
                norm
            ))
    print(f'[{time.time()-t0:5.1f}s] 클러스터링 대상 {len(posts)} / 빈 본문 제외 {skipped_empty} / 짧은 본문 제외 {skipped_short}')

    print(f'\n[{time.time()-t0:5.1f}s] === STEP 3: MinHash 생성 ===')
    minhashes = []
    sh_cache = []
    for i, p in enumerate(posts):
        sh = shingles(p[5])
        sh_cache.append(sh)
        m = MinHash(num_perm=NUM_PERM)
        for s in sh:
            m.update(s.encode('utf-8'))
        minhashes.append(m)
        if (i+1) % 5000 == 0:
            print(f'[{time.time()-t0:5.1f}s]   {i+1}/{len(posts)}')

    print(f'\n[{time.time()-t0:5.1f}s] === STEP 4: LSH 인덱스 + 후보 추출 ===')
    lsh = MinHashLSH(threshold=JACCARD_THRESHOLD, num_perm=NUM_PERM)
    for i, m in enumerate(minhashes):
        lsh.insert(i, m)

    pairs = []
    seen = set()
    for i, m in enumerate(minhashes):
        for j in lsh.query(m):
            if i >= j: continue
            if (i,j) in seen: continue
            seen.add((i,j))
            jsim = jaccard(sh_cache[i], sh_cache[j])
            if jsim >= JACCARD_THRESHOLD:
                pairs.append((i, j, jsim))
    print(f'[{time.time()-t0:5.1f}s] 유사 쌍: {len(pairs)}')

    parent = list(range(len(posts)))
    def find(x):
        while parent[x] != x: parent[x] = parent[parent[x]]; x = parent[x]
        return x
    def union(a,b):
        ra, rb = find(a), find(b)
        if ra != rb: parent[ra] = rb
    for i, j, _ in pairs: union(i,j)
    clusters = defaultdict(list)
    for i in range(len(posts)): clusters[find(i)].append(i)
    dup_clusters = {k:v for k,v in clusters.items() if len(v) > 1}
    print(f'[{time.time()-t0:5.1f}s] 클러스터 {len(dup_clusters)}, 영향받는 글 {sum(len(v) for v in dup_clusters.values())}')

    sim_bins = defaultdict(int)
    for _,_,s in pairs:
        if s >= 0.95: sim_bins['0.95-1.00'] += 1
        elif s >= 0.85: sim_bins['0.85-0.95'] += 1
        else: sim_bins['0.75-0.85'] += 1
    print('유사도 분포:', dict(sim_bins))

    print(f'\n[{time.time()-t0:5.1f}s] === STEP 5: 최종 xlsx 저장 ===')
    wb = Workbook(); wb.remove(wb.active)
    hdr_font = Font(bold=True, color='FFFFFF')
    hdr_fill_blue = PatternFill('solid', fgColor='1F4E78')
    hdr_fill_red  = PatternFill('solid', fgColor='C00000')

    new_header = list(H) + ['classification']
    def add_class_sheet(name, rows_to_add, default_class=None):
        ws = wb.create_sheet(name)
        ws.append(new_header)
        for c in ws[1]:
            c.font = hdr_font; c.fill = hdr_fill_blue
            c.alignment = Alignment(vertical='center')
        ws.freeze_panes = 'A2'
        for r in rows_to_add:
            tpl = r[TPL]
            cls = default_class or ('human' if (tpl is None or str(tpl).strip()=='') else 'ai_generated')
            ws.append(list(r) + [cls])
        for ci, cn in enumerate(new_header, 1):
            w = {'title':60,'url':50,'image':50,'ogImage':50,'metaDescription':50,
                 'ogDescription':50,'metaKeywords':40,'hashTags':40,'summary500':80,
                 'naverBlogUrl':40}.get(cn, 18)
            ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = w

    add_class_sheet('Human', human, 'human')
    add_class_sheet('AI_generated', ai, 'ai_generated')
    add_class_sheet('All', data)

    # Duplicate_Clusters
    ws_c = wb.create_sheet('Duplicate_Clusters')
    ws_c.append(['cluster_id','cluster_size','id','title','url','templateLabel','createdAt'])
    for c in ws_c[1]: c.font = hdr_font; c.fill = hdr_fill_red
    ws_c.freeze_panes = 'A2'
    cid = 0
    for root, members in sorted(dup_clusters.items(), key=lambda x: -len(x[1])):
        cid += 1
        for mi in members:
            p = posts[mi]
            ws_c.append([cid, len(members), p[0], p[1], p[2], p[3], p[4]])
    for col, w in [('A',10),('B',12),('C',22),('D',60),('E',55),('F',30),('G',25)]:
        ws_c.column_dimensions[col].width = w

    # Duplicate_Pairs
    ws_p = wb.create_sheet('Duplicate_Pairs')
    ws_p.append(['jaccard','id_a','title_a','url_a','id_b','title_b','url_b','templateLabel_a','templateLabel_b','createdAt_a','createdAt_b'])
    for c in ws_p[1]: c.font = hdr_font; c.fill = hdr_fill_red
    ws_p.freeze_panes = 'A2'
    for i,j,sim in sorted(pairs, key=lambda x:-x[2]):
        pa, pb = posts[i], posts[j]
        ws_p.append([round(sim,4), pa[0], pa[1], pa[2], pb[0], pb[1], pb[2], pa[3], pb[3], pa[4], pb[4]])
    for col, w in [('A',10),('B',22),('C',55),('D',50),('E',22),('F',55),('G',50),('H',28),('I',28),('J',22),('K',22)]:
        ws_p.column_dimensions[col].width = w

    # 시트 순서
    order = ['Human','AI_generated','All','Duplicate_Clusters','Duplicate_Pairs']
    for i, name in enumerate(order):
        if name in wb.sheetnames:
            cur = wb.sheetnames.index(name)
            wb.move_sheet(name, offset=i - cur)

    wb.save(DST)
    import os
    print(f'[{time.time()-t0:5.1f}s] 저장 완료: {DST} ({os.path.getsize(DST)/1024/1024:.1f} MB)')
    print(f'  시트: {wb.sheetnames}')
    print(f'\n=== 최종 요약 ===')
    print(f'  Human: {len(human)} / AI: {len(ai)} / All: {len(data)}')
    print(f'  유사 쌍(>=0.75): {len(pairs)} / 영향받는 글: {sum(len(v) for v in dup_clusters.values())}')
    print(f'  유사도 분포: {dict(sim_bins)}')

if __name__ == '__main__':
    main()
