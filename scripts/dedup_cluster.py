"""
MinHash + LSH 기반 유사 본문 클러스터링.
- 입력: drivingteacher_blog_posts_lite.xlsx (summary500 컬럼)
- 출력: 기존 classified.xlsx에 Duplicate_Clusters / Duplicate_Pairs 시트 추가
"""
import openpyxl
from openpyxl.styles import Font, PatternFill
from datasketch import MinHash, MinHashLSH
import re
import time
from collections import defaultdict

SRC = '/Users/simjaehyeong/Desktop/adrock/seo/data/drivingteacher_blog_posts_lite.xlsx'
DST = '/Users/simjaehyeong/Desktop/adrock/seo/data/drivingteacher_blog_posts_classified.xlsx'

SHINGLE_K = 5            # 문자 n-gram 크기
NUM_PERM = 128           # MinHash 해상도
JACCARD_THRESHOLD = 0.75 # LSH 후보 임계치

def normalize(text: str) -> str:
    text = text or ''
    text = re.sub(r'https?://\S+', ' ', text)   # URL 제거
    text = re.sub(r'[^\w가-힣]+', ' ', text)     # 특수문자 제거
    text = re.sub(r'\s+', ' ', text).strip()
    return text

def shingles(text: str, k: int = SHINGLE_K):
    text = normalize(text)
    if len(text) < k:
        return set([text])
    return {text[i:i+k] for i in range(len(text) - k + 1)}

def jaccard(a: set, b: set) -> float:
    if not a or not b: return 0.0
    inter = len(a & b)
    union = len(a | b)
    return inter / union if union else 0.0

def main():
    t0 = time.time()
    print(f'[{time.time()-t0:5.1f}s] 원본 로드...')
    wb_src = openpyxl.load_workbook(SRC, read_only=True)
    ws = wb_src['Posts']
    rows = list(ws.iter_rows(values_only=True))
    H = rows[0]; data = rows[1:]
    idx = {c:i for i,c in enumerate(H)}
    print(f'[{time.time()-t0:5.1f}s]  -> {len(data)}건')

    # MinHash 생성
    print(f'[{time.time()-t0:5.1f}s] MinHash 생성 (k={SHINGLE_K}, perm={NUM_PERM})...')
    minhashes = []
    shingle_cache = []
    for i, r in enumerate(data):
        text = r[idx['summary500']] or ''
        sh = shingles(text)
        shingle_cache.append(sh)
        m = MinHash(num_perm=NUM_PERM)
        for s in sh:
            m.update(s.encode('utf-8'))
        minhashes.append(m)
        if (i+1) % 5000 == 0:
            print(f'[{time.time()-t0:5.1f}s]   {i+1}/{len(data)}')

    # LSH 인덱스
    print(f'[{time.time()-t0:5.1f}s] LSH 인덱스 구축 (threshold={JACCARD_THRESHOLD})...')
    lsh = MinHashLSH(threshold=JACCARD_THRESHOLD, num_perm=NUM_PERM)
    for i, m in enumerate(minhashes):
        lsh.insert(i, m)

    # 후보 쌍 추출 + 실제 Jaccard 검증
    print(f'[{time.time()-t0:5.1f}s] 후보 쌍 추출...')
    pairs = []
    seen = set()
    for i, m in enumerate(minhashes):
        cands = lsh.query(m)
        for j in cands:
            if i >= j: continue
            key = (i, j)
            if key in seen: continue
            seen.add(key)
            # 실제 Jaccard로 검증
            j_sim = jaccard(shingle_cache[i], shingle_cache[j])
            if j_sim >= JACCARD_THRESHOLD:
                pairs.append((i, j, j_sim))
        if (i+1) % 5000 == 0:
            print(f'[{time.time()-t0:5.1f}s]   {i+1}/{len(data)}, 후보 쌍: {len(pairs)}')
    print(f'[{time.time()-t0:5.1f}s] 유사 쌍 총: {len(pairs)}')

    # Union-Find로 클러스터링
    print(f'[{time.time()-t0:5.1f}s] 클러스터링...')
    parent = list(range(len(data)))
    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x
    def union(a,b):
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[ra] = rb
    for i, j, _ in pairs:
        union(i, j)
    clusters = defaultdict(list)
    for i in range(len(data)):
        clusters[find(i)].append(i)
    dup_clusters = {k:v for k,v in clusters.items() if len(v) > 1}
    print(f'[{time.time()-t0:5.1f}s] 클러스터(>=2): {len(dup_clusters)}, 영향받는 글: {sum(len(v) for v in dup_clusters.values())}')

    # 분포
    size_dist = defaultdict(int)
    for v in dup_clusters.values():
        size_dist[len(v)] += 1
    print('클러스터 크기 분포:')
    for sz in sorted(size_dist):
        print(f'   size {sz}: {size_dist[sz]}건')

    # 출력: 기존 classified xlsx에 시트 추가
    print(f'[{time.time()-t0:5.1f}s] xlsx 저장...')
    wb = openpyxl.load_workbook(DST)
    for sheet_name in ['Duplicate_Clusters', 'Duplicate_Pairs']:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    hdr_font = Font(bold=True, color='FFFFFF')
    hdr_fill = PatternFill('solid', fgColor='C00000')

    # 클러스터 시트
    ws_c = wb.create_sheet('Duplicate_Clusters')
    ws_c.append(['cluster_id','cluster_size','member_idx','id','title','url','templateLabel','createdAt','summary500_head'])
    for c in ws_c[1]:
        c.font = hdr_font; c.fill = hdr_fill
    ws_c.freeze_panes = 'A2'
    cid = 0
    for root, members in sorted(dup_clusters.items(), key=lambda x: -len(x[1])):
        cid += 1
        for m_idx in members:
            r = data[m_idx]
            ws_c.append([
                cid, len(members), m_idx,
                r[idx['id']], r[idx['title']], r[idx['url']],
                r[idx['templateLabel']], r[idx['createdAt']],
                (r[idx['summary500']] or '')[:200],
            ])
    # 너비
    widths = {'A':10,'B':12,'C':10,'D':22,'E':50,'F':50,'G':25,'H':25,'I':80}
    for col, w in widths.items():
        ws_c.column_dimensions[col].width = w

    # 쌍 시트
    ws_p = wb.create_sheet('Duplicate_Pairs')
    ws_p.append(['jaccard','id_a','title_a','url_a','id_b','title_b','url_b','templateLabel_a','templateLabel_b','createdAt_a','createdAt_b'])
    for c in ws_p[1]:
        c.font = hdr_font; c.fill = hdr_fill
    ws_p.freeze_panes = 'A2'
    for i, j, sim in sorted(pairs, key=lambda x:-x[2]):
        ra, rb = data[i], data[j]
        ws_p.append([
            round(sim,4),
            ra[idx['id']], ra[idx['title']], ra[idx['url']],
            rb[idx['id']], rb[idx['title']], rb[idx['url']],
            ra[idx['templateLabel']], rb[idx['templateLabel']],
            ra[idx['createdAt']], rb[idx['createdAt']],
        ])
    widths_p = {'A':10,'B':22,'C':50,'D':50,'E':22,'F':50,'G':50,'H':25,'I':25,'J':25,'K':25}
    for col, w in widths_p.items():
        ws_p.column_dimensions[col].width = w

    wb.save(DST)
    print(f'[{time.time()-t0:5.1f}s] 완료: {DST}')
    print(f'  - 시트: Duplicate_Clusters ({sum(len(v) for v in dup_clusters.values())}행), Duplicate_Pairs ({len(pairs)}행)')

if __name__ == '__main__':
    main()
